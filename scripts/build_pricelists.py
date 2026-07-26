"""Maintainer-side ERU XLSX -> bundled pricelist JSON compiler.

Output JSON schema (v1):

{
  "schema_version": 1,
  "generated_at": "2026-07-19T00:00:00Z",
  "year": 2026,
  "sources": {
    "<source_file_name>": {
      "source_url": "file:///abs/path/price_file.xlsx",
      "fetched_at": "2026-07-19T00:00:00Z",
      "sha256": "..."
    }
  },
  "distributors": {
    "cez": {
      "VT": {"unit": "Kc/A/mesic", "price_incl_vat": 12.34, "price_excl_vat": 9.5},
      "NT": {"unit": "Kc/A/mesic", "price_incl_vat": 9.1, "price_excl_vat": 7.4},
      "POZE": {"unit": "Kc/A/mesic", "price_incl_vat": 2.1, "price_excl_vat": 1.6}
    },
    "egd": { ... },
    "pre": { ... }
  },
  "valid_from_snapshots": [
    {
      "valid_from": "2026-01-01",
      "distributors": {
        "cez": {
          "VT": {"unit": "Kc/A/mesic", "price_incl_vat": 12.34, "price_excl_vat": 9.5},
          "NT": {"unit": "Kc/A/mesic", "price_incl_vat": 9.1, "price_excl_vat": 7.4},
          "POZE": {"unit": "Kc/A/mesic", "price_incl_vat": 2.1, "price_excl_vat": 1.6}
        },
        "egd": { ... },
        "pre": { ... }
      },
      "sources": {
        "cez": "price_file.xlsx",
        "egd": "price_file.xlsx",
        "pre": "price_file.xlsx"
      }
    }
  ]
}

ERU-decree mode (schema v2, --eru-decree / auto-detected):

The real published NN price decree (e.g. ceny-nn26-1.xlsx) is not a flat table; its
`Distribuce` sheet is a sequence of per-sazba blocks. In this mode the output is:

{
  "schema_version": 2,
  "year": 2026,
  "vat_rate": 0.21,
  "sources": {"ceny-nn26-1.xlsx": {"source_url": "https://eru.gov.cz/...", ...}},
  "distributors": {
    "cez": {
      "D01d": {"unit": "Kc/MWh", "price_incl_vat": ..., "price_excl_vat": ...,
               "vt": {...}},                       # single-tariff: VT only
      "D25d": {"unit": "Kc/MWh", "price_incl_vat": ...(VT), "price_excl_vat": ...(VT),
               "vt": {...}, "nt": {...}},           # two-tariff: VT high + NT low
      "POZE": {"unit": "Kc/A/mesic", "price_incl_vat": 0.0, "price_excl_vat": 0.0}
    }, "egd": {...}, "pre": {...}
  },
  "valid_from_snapshots": [{"valid_from": "2026-01-01", "distributors": {...}, "sources": {...}}]
}

The tariff's top-level price fields mirror the VT leg so the existing 2-level readers
(distributor -> tariff -> {unit, price_incl_vat, price_excl_vat}) keep working; the `vt`/`nt`
sub-objects are the canonical per-leg data. Only POZE carries "Kc/A/mesic"; VT/NT are "Kc/MWh".

Reader behavior (flat fixture mode, schema v1):
- Sheets are discovered by header text, not by fixed index.
- Price rows are resolved by column headers and converted as:
    - bold cell -> price_incl_vat
    - parenthesised string "(...) -> price_excl_vat
- Unit must be "Kc/A/mesic" (never "Kc/MWh").
- Tariff coverage must be complete for every required distributor (cez, egd, pre)
  and every discovered tariff in each snapshot.
- Any per-tariff price movement > 30% vs previous valid_from snapshot causes failure,
  unless --allow-large-moves is supplied.
- On any validation/build failure, output path is never created or modified.
"""

from __future__ import annotations

import argparse
import hashlib
import json
import re
import sys
import tempfile
import unicodedata
from dataclasses import dataclass
from datetime import datetime, date, timezone
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
from urllib.parse import quote
import os

from openpyxl import load_workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet

REQUIRED_DSOS = {"cez", "egd", "pre"}
TARIFF_PRICE_FIELDS = ("price_incl_vat", "price_excl_vat")
REQUIRED_HEADER_COLUMNS = ("tariff", "unit", "price", "valid_from")

HEADER_PATTERNS: Dict[str, Tuple[str, ...]] = {
    "distributor": ("distributor", "provider", "dso", "supplier"),
    "tariff": ("tariff", "tarifa", "sazba"),
    "unit": ("unit", "jednotka"),
    "price": ("price", "cena", "amount", "value"),
    "valid_from": (
        "valid",
        "platna_od",
        "platnost",
        "platnost_od",
        "as_of",
        "date",
    ),
}


def _parse_args(argv: Optional[List[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Build bundled pricelists JSON from ERU XLSX files.")
    parser.add_argument(
        "--output",
        required=True,
        type=Path,
        help="Output JSON path",
    )
    parser.add_argument(
        "--allow-large-moves",
        action="store_true",
        help=(
            "Allow price changes greater than 30% against previous snapshot. "
            "Required only for intentional tariff changes."
        ),
    )
    parser.add_argument(
        "--source-url",
        action="append",
        default=[],
        metavar="FILENAME=URL",
        help=(
            "Real https provenance URL for a source file, as 'filename=https://...'. "
            "REQUIRED for every file in ERU-decree mode so the shipped JSON never carries a "
            "file:// URL. Repeatable."
        ),
    )
    parser.add_argument(
        "--valid-from",
        default="2026-01-01",
        help="valid_from date (ISO-8601) for the snapshot built in ERU-decree mode.",
    )
    parser.add_argument(
        "--vat-rate",
        type=float,
        default=0.21,
        help="Statutory VAT rate used to derive price_incl_vat from the excl-VAT decree prices.",
    )
    parser.add_argument(
        "--eru-decree",
        action="store_true",
        help=(
            "Force ERU price-decree layout (Distribuce sheet with per-tariff D-sazby blocks). "
            "When omitted the layout is auto-detected; the flat fixture layout is used otherwise."
        ),
    )
    parser.add_argument("files", nargs="+", type=Path, help="Versioned ERU XLSX source files")
    return parser.parse_args(argv)


def _parse_source_url_map(pairs: List[str]) -> Dict[str, str]:
    mapping: Dict[str, str] = {}
    for pair in pairs:
        if "=" not in pair:
            raise BuildError(f"--source-url must be 'filename=url', got {pair!r}")
        name, url = pair.split("=", 1)
        name, url = name.strip(), url.strip()
        if not name or not url:
            raise BuildError(f"--source-url must be 'filename=url', got {pair!r}")
        if not url.startswith("https://"):
            raise BuildError(f"--source-url for {name!r} must be an https URL, got {url!r}")
        mapping[name] = url
    return mapping


class BuildError(RuntimeError):
    """Domain-specific build error raised for non-zero exit conditions."""


@dataclass(frozen=True)
class SourceMeta:
    source_file: str
    source_url: str
    fetched_at: str
    sha256: str


@dataclass
class PricePoint:
    unit: str
    price_incl_vat: Optional[float] = None
    price_excl_vat: Optional[float] = None
    source_file: Optional[str] = None

    def to_payload(self) -> Dict[str, Any]:
        return {
            "unit": self.unit,
            "price_incl_vat": self.price_incl_vat,
            "price_excl_vat": self.price_excl_vat,
        }


def _normalize_text(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", "", str(value).strip().lower())


def _normalize_dso(value: Any) -> str:
    normalized = _normalize_text(value)
    if not normalized:
        return ""
    if "cez" in normalized:
        return "cez"
    if "egd" in normalized:
        return "egd"
    if "pre" in normalized:
        return "pre"
    return ""


def _normalize_unit(raw: Any) -> str:
    normalized = re.sub(r"\s+", "", str(raw)).replace("Kč", "Kc").replace("měsíc", "mesic")
    return normalized


def _normalize_header_token(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"[^a-z]", "", str(value).strip().lower())


def _parse_date(raw: Any) -> str:
    if isinstance(raw, datetime):
        return raw.date().isoformat()
    if isinstance(raw, date):
        return raw.isoformat()
    if isinstance(raw, int):
        raise BuildError(f"valid_from value {raw!r} is invalid date format")

    if raw is None:
        raise BuildError("valid_from value is missing")

    text = str(raw).strip()
    if not text:
        raise BuildError("valid_from value is missing")

    if re.fullmatch(r"\d{4}-\d{2}-\d{2}", text):
        return text
    raise BuildError(f"valid_from value {raw!r} is not ISO-8601 date")


def _parse_price(cell: Cell) -> Tuple[Optional[str], float]:
    raw = cell.value
    if raw is None:
        raise BuildError("price cell is empty")

    text = str(raw).strip()
    is_excl = False
    if text.startswith("(") and text.endswith(")"):
        is_excl = True
        text = text[1:-1].strip()

    if not text:
        raise BuildError("price cell is empty")

    try:
        price = float(text.replace(",", "."))
    except ValueError as exc:
        raise BuildError(f"price cell {raw!r} is not numeric") from exc

    is_incl = bool(cell.font and cell.font.bold)
    if cell.number_format and "(" in str(cell.number_format) and ")" in str(cell.number_format):
        if not is_excl:
            is_excl = True
        is_incl = False

    if is_excl == is_incl:
        raise BuildError("price cell is neither explicitly incl nor excl via bold/parentheses")

    return ("price_excl_vat" if is_excl else "price_incl_vat"), price


def _find_header_columns(worksheet: Worksheet) -> Tuple[Optional[Dict[str, int]], Optional[int]]:
    max_row = min(worksheet.max_row or 1, 40)
    max_column = worksheet.max_column or 1

    for row_idx in range(1, max_row + 1):
        row = list(worksheet.iter_rows(min_row=row_idx, max_row=row_idx, max_col=max_column, values_only=False))[0]
        column_map: Dict[str, int] = {}
        for column_idx, cell in enumerate(row, start=1):
            token = _normalize_header_token(cell.value)
            if not token:
                continue
            for key, aliases in HEADER_PATTERNS.items():
                if any(alias in token for alias in aliases) and key not in column_map:
                    column_map[key] = column_idx

        missing = [name for name in REQUIRED_HEADER_COLUMNS if name not in column_map]
        if not missing:
            return column_map, row_idx
    return None, None


def _sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1 << 20), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _read_source_metadata(path: Path) -> SourceMeta:
    stat = path.stat()
    fetched = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat()
    return SourceMeta(
        source_file=path.name,
        source_url=f"file://{quote(str(path.resolve()))}",
        fetched_at=fetched,
        sha256=_sha256(path),
    )


def _build_snapshot_map(
    files: List[Path],
) -> Tuple[
    Dict[str, Dict[str, Dict[str, PricePoint]]],
    Dict[str, str],
    Dict[str, SourceMeta],
]:
    source_files: Dict[str, SourceMeta] = {}
    # valid_from -> distributor -> tariff -> price point
    snapshots: Dict[str, Dict[str, Dict[str, PricePoint]]] = {}
    # valid_from -> distributor -> source_file
    snapshot_sources: Dict[str, str] = {}

    for path in files:
        if not path.exists():
            raise BuildError(f"input file does not exist: {path}")
        metadata = _read_source_metadata(path)
        source_files[metadata.source_file] = metadata

        workbook = load_workbook(path, data_only=True)
        if not workbook.worksheets:
            raise BuildError(f"input file has no sheets: {path}")

        for worksheet in workbook.worksheets:
            header_columns, header_row = _find_header_columns(worksheet)
            if not header_columns:
                continue

            for row in worksheet.iter_rows(min_row=header_row + 1, values_only=False):
                if all(cell.value in (None, "") for cell in row):
                    continue
                dso_raw = row[header_columns["distributor"] - 1].value if "distributor" in header_columns else None
                dso = _normalize_dso(dso_raw or worksheet.title)
                if not dso:
                    raise BuildError(f"could not resolve distributor in {worksheet.title}")
                if dso not in REQUIRED_DSOS:
                    raise BuildError(f"unexpected distributor {dso_raw!r} in {worksheet.title}")

                tariff_raw = row[header_columns["tariff"] - 1].value
                if tariff_raw in (None, ""):
                    continue
                tariff = str(tariff_raw).strip()
                if not tariff:
                    continue

                unit_raw = row[header_columns["unit"] - 1].value
                unit = _normalize_unit(unit_raw)
                if unit != "Kc/A/mesic":
                    raise BuildError(f"unsupported unit {unit!r} for {dso}/{tariff}")

                valid_from_raw = row[header_columns["valid_from"] - 1].value
                valid_from = _parse_date(valid_from_raw)

                price_cell = row[header_columns["price"] - 1]
                price_key, price = _parse_price(price_cell)

                snapshot = snapshots.setdefault(valid_from, {})
                snapshot.setdefault(dso, {})
                point = snapshot[dso].setdefault(tariff, PricePoint(unit=unit, source_file=metadata.source_file))
                if getattr(point, price_key) is not None:
                    raise BuildError(f"duplicate {price_key} for {dso}/{tariff}/{valid_from}")
                setattr(point, price_key, price)
                snapshot_sources.setdefault(valid_from, metadata.source_file)

            # no row-level break; multiple sheets can contribute to different distributors
        workbook.close()

    if not snapshots:
        raise BuildError("no valid price rows were parsed from any worksheet")
    return snapshots, snapshot_sources, source_files


def _snapshot_coverage_complete(
    snapshots: Dict[str, Dict[str, Dict[str, PricePoint]]],
) -> None:
    for valid_from, distributors in snapshots.items():
        missing_dsos = REQUIRED_DSOS - distributors.keys()
        if missing_dsos:
            raise BuildError(
                f"snapshot {valid_from} misses required distributor(s): {", ".join(sorted(missing_dsos))}"
            )

        tariff_union = set()
        for dso_data in distributors.values():
            tariff_union.update(dso_data.keys())
        if not tariff_union:
            raise BuildError(f"snapshot {valid_from} contains no tariff rows")

        for distributor, dso_data in distributors.items():
            missing_tariffs = tariff_union - dso_data.keys()
            if missing_tariffs:
                raise BuildError(
                    f"{distributor} snapshot {valid_from} missing tariff(s): {", ".join(sorted(missing_tariffs))}"
                )

            for tariff, point in dso_data.items():
                if point.unit != "Kc/A/mesic":
                    raise BuildError(
                        f"{distributor} snapshot {valid_from} {tariff} unit must be Kc/A/mesic"
                    )
                for field in TARIFF_PRICE_FIELDS:
                    if getattr(point, field) is None:
                        raise BuildError(
                            f"{distributor} snapshot {valid_from} {tariff} missing field {field}"
                        )


def _check_price_movement(
    snapshots: Dict[str, Dict[str, Dict[str, PricePoint]]], allow_large_moves: bool
) -> None:
    if allow_large_moves:
        return

    sorted_dates = sorted(snapshots.keys())
    if len(sorted_dates) <= 1:
        return

    for prev_date, cur_date in zip(sorted_dates[:-1], sorted_dates[1:]):
        prev = snapshots[prev_date]
        cur = snapshots[cur_date]
        for distributor in sorted(REQUIRED_DSOS):
            for tariff in sorted(prev[distributor].keys()):
                prev_point = prev[distributor][tariff]
                cur_point = cur[distributor][tariff]
                for field in TARIFF_PRICE_FIELDS:
                    previous = getattr(prev_point, field)
                    current = getattr(cur_point, field)
                    if previous is None or current is None:
                        continue
                    if previous == 0:
                        raise BuildError(
                            f"{distributor}/{tariff} {field} moved from zero to {current} without override"
                        )
                    if abs(current - previous) / previous > 0.3:
                        raise BuildError(
                            f"{distributor}/{tariff} {field} moved >30% from {previous} on {prev_date} to {current} on {cur_date}"
                        )


def _build_payload(snapshots: Dict[str, Dict[str, Dict[str, PricePoint]]], source_files: Dict[str, SourceMeta]) -> Dict[str, Any]:
    sorted_dates = sorted(snapshots.keys())
    latest = snapshots[sorted_dates[-1]]

    def convert_distributor_data(data: Dict[str, Dict[str, PricePoint]]) -> Dict[str, Any]:
        result: Dict[str, Any] = {}
        for tariff, point in sorted(data.items()):
            result[tariff] = point.to_payload()
        return result

    valid_from_snapshots = []
    for valid_from in sorted_dates:
        snapshot_distributors = snapshots[valid_from]
        valid_from_snapshots.append(
            {
                "valid_from": valid_from,
                "distributors": {
                    distributor: convert_distributor_data(snapshot_distributors[distributor])
                    for distributor in sorted(snapshot_distributors)
                },
                "sources": {
                    distributor: snapshot_distributors[distributor][next(iter(snapshot_distributors[distributor]))].source_file
                    for distributor in sorted(snapshot_distributors)
                },
            }
        )

    year = int(sorted_dates[-1][:4])
    payload = {
        "schema_version": 1,
        "generated_at": datetime.now(tz=timezone.utc).replace(microsecond=0).isoformat(),
        "year": year,
        "sources": {
            source_file: {
                "source_url": data.source_url,
                "fetched_at": data.fetched_at,
                "sha256": data.sha256,
            }
            for source_file, data in sorted(source_files.items(), key=lambda item: item[0])
        },
        "distributors": {distributor: convert_distributor_data(latest[distributor]) for distributor in sorted(latest)},
        "valid_from_snapshots": valid_from_snapshots,
    }
    return payload


def _write_payload(path: Path, payload: Dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with tempfile.NamedTemporaryFile("w", delete=False, dir=path.parent, suffix=".tmp", encoding="utf-8") as tmp:
        json.dump(payload, tmp, ensure_ascii=False, indent=2, sort_keys=True)
        tmp.flush()
        temp_path = Path(tmp.name)
    os.replace(temp_path, path)


# ---------------------------------------------------------------------------
# ERU price-decree layout (real published XLSX, e.g. ceny-nn26-1.xlsx)
#
# The regulator's real "Distribuce" sheet is NOT a flat tariff/unit/price table.
# It is a sequence of per-sazba blocks:
#
#     Sazba D 25d - Dvoutarifová ...
#       ... breaker capacity table (Kč/měsíc) ...
#       · z platu za distribuované množství elektřiny ve vysokém tarifu:
#           ČEZ   EG.D  PRE
#       Kč/MWh  2252.45 2243.88 1656.49        <- VT energy price (excl VAT)
#       · z platu za distribuované množství elektřiny v nízkém tarifu:
#           ČEZ   EG.D  PRE
#       Kč/MWh   116.5  224.3  175.2           <- NT energy price (excl VAT)
#
# Single-tariff sazby (D01d, D02d) carry a single "... distribuované množství
# elektřiny:" section (VT only, no NT). Sheets/sections are located by Czech text,
# never by fixed offsets (formats drift between years). Decree prices are WITHOUT
# VAT (the sheet footnote states they exclude "daň z elektřiny a DPH"), so
# price_incl_vat is derived as excl * (1 + vat_rate). The POZE household component
# ("na podporu ... podporovaných zdrojů energie", per rezervovaný příkon) lives on
# the "Regulovaná složka NN" sheet in Kč/A/měsíc — the only per-ampere item.
# ---------------------------------------------------------------------------

ERU_DSO_HEADERS = {
    "cez": "cez",
    "egd": "egd",
    "eg.d": "egd",
    "eg d": "egd",
    "pre": "pre",
}
_ERU_TARIFF_RE = re.compile(r"sazba\s+([cd])\s*(\d+d)\b")
# The tariff-header row's own text carries a short description after the code
# (e.g. "Sazba D 25d - Dvoutarifová sazba s operativním řízením ..."), one
# per household D-tariff, identical across all three distributors — matched
# on the RAW (accented) cell, unlike `_ERU_TARIFF_RE` which runs against the
# accent-stripped routing text.
_ERU_DESC_PREFIX_RE = re.compile(r"^sazba\s+[a-z]\s*\d+d\s*[-–]\s*(.+)$", re.IGNORECASE)
_ERU_VT_HINT = "distribuovane mnozstvi elektriny ve vysokem tarifu"
_ERU_NT_HINT = "distribuovane mnozstvi elektriny v nizkem tarifu"
_ERU_SINGLE_HINT = "distribuovane mnozstvi elektriny:"


def _strip_accents(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(ch for ch in text if not unicodedata.combining(ch))
    return re.sub(r"\s+", " ", text).strip().lower()


def _eru_dso_of(value: Any) -> Optional[str]:
    return ERU_DSO_HEADERS.get(_strip_accents(value).replace(" ", ""))


def _eru_description(raw: Any) -> str:
    """Description text after "Sazba <cat> <code> - " on the header row, with
    diacritics preserved (unlike the accent-stripped routing match)."""
    text = "" if raw is None else str(raw)
    text = re.sub(r"\s+", " ", text.replace("\xa0", " ")).strip()
    match = _ERU_DESC_PREFIX_RE.match(text)
    return match.group(1).strip() if match else text


def _is_eru_workbook(workbook) -> bool:
    for worksheet in workbook.worksheets:
        if "distribuce" in _strip_accents(worksheet.title):
            for row in worksheet.iter_rows(min_row=1, max_row=min(worksheet.max_row or 1, 60), values_only=True):
                if row and _ERU_TARIFF_RE.match(_strip_accents(row[0])):
                    return True
    return False


def _eru_price_row(rows: List[List[Any]], start: int) -> Optional[Dict[str, float]]:
    """From a section-label row, locate the DSO header row then the Kč/MWh value row."""
    header: Optional[Dict[str, int]] = None
    header_idx = -1
    for r in range(start, min(start + 4, len(rows))):
        cols: Dict[str, int] = {}
        for ci, value in enumerate(rows[r]):
            dso = _eru_dso_of(value)
            if dso and dso not in cols:
                cols[dso] = ci
        if REQUIRED_DSOS <= cols.keys():
            header, header_idx = cols, r
            break
    if header is None:
        return None
    for r in range(header_idx + 1, min(header_idx + 3, len(rows))):
        if _strip_accents(rows[r][0]).startswith("kc/mwh"):
            result: Dict[str, float] = {}
            for dso, ci in header.items():
                raw = rows[r][ci]
                if not isinstance(raw, (int, float)):
                    raise BuildError(f"non-numeric Kč/MWh value {raw!r} for {dso} at row {r + 1}")
                result[dso] = float(raw)
            return result
    raise BuildError(f"could not locate Kč/MWh value row after section at row {start + 1}")


def _extract_eru_poze(workbook) -> Dict[str, float]:
    for worksheet in workbook.worksheets:
        if "regulovana slozka" not in _strip_accents(worksheet.title):
            continue
        for row in worksheet.iter_rows(values_only=True):
            for ci, cell in enumerate(row):
                label = _strip_accents(cell)
                if "na podporu elektriny z podporovanych zdroju" in label:
                    value = row[ci + 1] if ci + 1 < len(row) else None
                    unit = _strip_accents(row[ci + 2]) if ci + 2 < len(row) else ""
                    if not isinstance(value, (int, float)):
                        raise BuildError(f"POZE value is not numeric: {value!r}")
                    if "kc/a" not in unit:
                        raise BuildError(f"POZE unit must be per-ampere (Kč/A/měsíc), got {row[ci + 2]!r}")
                    return {"excl": float(value)}
    raise BuildError("could not find POZE (podpora podporovaných zdrojů) row on 'Regulovaná složka' sheet")


def _parse_eru_tariffs(
    worksheet,
) -> Tuple[Dict[str, Dict[str, Dict[str, float]]], Dict[str, str]]:
    rows = [
        [worksheet.cell(r, c).value for c in range(1, (worksheet.max_column or 1) + 1)]
        for r in range(1, (worksheet.max_row or 1) + 1)
    ]
    tariffs: Dict[str, Dict[str, Dict[str, float]]] = {}
    descriptions: Dict[str, str] = {}
    current: Optional[str] = None
    for i, row in enumerate(rows):
        raw = row[0]
        label = _strip_accents(raw)
        match = _ERU_TARIFF_RE.match(label)
        if match:
            # Household tariffs are category D; category C (business) is skipped.
            current = f"D{match.group(2)}" if match.group(1) == "d" else None
            if current:
                tariffs.setdefault(current, {})
                descriptions[current] = _eru_description(raw)
            continue
        if current is None:
            continue
        if _ERU_VT_HINT in label:
            tariffs[current]["VT"] = _eru_price_row(rows, i)
        elif _ERU_NT_HINT in label:
            tariffs[current]["NT"] = _eru_price_row(rows, i)
        elif _ERU_SINGLE_HINT in label and "VT" not in tariffs[current]:
            tariffs[current]["VT"] = _eru_price_row(rows, i)
    if not tariffs:
        raise BuildError("no D-category tariff blocks found on Distribuce sheet")
    return tariffs, descriptions


def _round2(value: float) -> float:
    return round(value + 1e-9, 2)


def _eru_point(unit: str, excl: float, vat_rate: float) -> Dict[str, Any]:
    return {
        "unit": unit,
        "price_excl_vat": _round2(excl),
        "price_incl_vat": _round2(excl * (1.0 + vat_rate)),
    }


def _eru_tariff_payload(
    legs: Dict[str, Dict[str, float]], dso: str, vat_rate: float, description: str
) -> Dict[str, Any]:
    """One D-tariff rate for a distributor.

    Shape is 2-level-reader compatible: the tariff's top-level price fields mirror
    the VT (high/single) leg so the existing config_registry / pricelists-endpoint
    readers (distributor -> tariff -> {unit, price_incl_vat, price_excl_vat}) keep
    working. The canonical per-leg data lives in the `vt` (and, for two-tariff
    sazby, `nt`) sub-objects, both in Kc/MWh. `description` is the ERU decree's
    own short-form CZ label for the sazba (owner UX rev item 2) — additive, no
    existing reader keys on it.
    """
    if "VT" not in legs or dso not in legs["VT"]:
        raise BuildError(f"tariff missing VT price for {dso}")
    vt = _eru_point("Kc/MWh", legs["VT"][dso], vat_rate)
    payload: Dict[str, Any] = {**vt, "vt": vt, "description": description}
    if "NT" in legs:
        if dso not in legs["NT"]:
            raise BuildError(f"two-tariff sazba missing NT price for {dso}")
        payload["nt"] = _eru_point("Kc/MWh", legs["NT"][dso], vat_rate)
    return payload


def _build_eru_distributors(
    tariffs: Dict[str, Dict[str, Dict[str, float]]],
    descriptions: Dict[str, str],
    poze: Dict[str, float],
    vat_rate: float,
) -> Dict[str, Any]:
    distributors: Dict[str, Any] = {}
    for dso in sorted(REQUIRED_DSOS):
        rates: Dict[str, Any] = {}
        for code in sorted(tariffs):
            rates[code] = _eru_tariff_payload(tariffs[code], dso, vat_rate, descriptions.get(code, ""))
        # POZE is the only per-ampere item; kept as a tariff-level key for
        # backward compatibility with the existing 2-level readers.
        rates["POZE"] = _eru_point("Kc/A/mesic", poze["excl"], vat_rate)
        distributors[dso] = rates
    return distributors


def _validate_eru_distributors(distributors: Dict[str, Any], required_codes: set) -> None:
    for dso in sorted(REQUIRED_DSOS):
        if dso not in distributors:
            raise BuildError(f"missing distributor {dso}")
        rates = distributors[dso]
        missing = required_codes - (rates.keys() - {"POZE"})
        if missing:
            raise BuildError(f"{dso} missing D-tariff(s): {', '.join(sorted(missing))}")
        for code in required_codes:
            rate = rates[code]
            if rate["unit"] != "Kc/MWh":
                raise BuildError(f"{dso}/{code} unit must be Kc/MWh, got {rate['unit']!r}")
            for leg in ("vt",) + (("nt",) if "nt" in rate else ()):
                point = rate[leg]
                if point["unit"] != "Kc/MWh":
                    raise BuildError(f"{dso}/{code}/{leg} unit must be Kc/MWh, got {point['unit']!r}")
                if point["price_incl_vat"] is None or point["price_excl_vat"] is None:
                    raise BuildError(f"{dso}/{code}/{leg} missing a price field")
        if rates["POZE"]["unit"] != "Kc/A/mesic":
            raise BuildError(f"{dso} POZE unit must be Kc/A/mesic")


def _build_eru(args: argparse.Namespace) -> Dict[str, Any]:
    source_url_map = _parse_source_url_map(args.source_url)
    if len(args.files) != 1:
        raise BuildError("ERU-decree mode expects exactly one source XLSX")
    path = args.files[0]
    if not path.exists():
        raise BuildError(f"input file does not exist: {path}")
    if path.name not in source_url_map:
        raise BuildError(
            f"ERU-decree mode requires --source-url {path.name}=https://... so the shipped "
            "JSON records the real regulator URL (never file://)"
        )

    valid_from = _parse_date(args.valid_from)
    workbook = load_workbook(path, data_only=True)
    try:
        distribuce = next(
            (ws for ws in workbook.worksheets if "distribuce" in _strip_accents(ws.title)), None
        )
        if distribuce is None:
            raise BuildError("ERU workbook has no 'Distribuce' sheet")
        tariffs, descriptions = _parse_eru_tariffs(distribuce)
        poze = _extract_eru_poze(workbook)
    finally:
        workbook.close()

    distributors = _build_eru_distributors(tariffs, descriptions, poze, args.vat_rate)
    _validate_eru_distributors(distributors, set(tariffs.keys()))

    stat = path.stat()
    fetched = datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).replace(microsecond=0).isoformat()
    source_meta = {
        "source_url": source_url_map[path.name],
        "fetched_at": fetched,
        "sha256": _sha256(path),
        "decree": "14/2025 (Energetický regulační věstník 18/2025), distribuce NN pro rok 2026",
        "note": (
            "POZE (příspěvek na podporované zdroje) je pro rok 2026 vynulován výměrem 15/2025 "
            "(navazuje na 13/2025), účinnost od 2026-01-01; zveřejněný XLSX už tuto nulu obsahuje. "
            "price_incl_vat je odvozeno z cen bez DPH v XLSX sazbou vat_rate."
        ),
    }
    snapshot = {
        "valid_from": valid_from,
        "distributors": distributors,
        "sources": {dso: path.name for dso in sorted(REQUIRED_DSOS)},
    }
    return {
        "schema_version": 2,
        "generated_at": datetime.now(tz=timezone.utc).replace(microsecond=0).isoformat(),
        "year": int(valid_from[:4]),
        "vat_rate": args.vat_rate,
        "sources": {path.name: source_meta},
        "distributors": distributors,
        "valid_from_snapshots": [snapshot],
    }


def _eru_mode_requested(args: argparse.Namespace) -> bool:
    if args.eru_decree:
        return True
    for path in args.files:
        if not path.exists():
            return False
        workbook = load_workbook(path, data_only=True)
        try:
            if _is_eru_workbook(workbook):
                return True
        finally:
            workbook.close()
    return False


def build(argv: Optional[List[str]] = None) -> int:
    args = _parse_args(argv)

    if _eru_mode_requested(args):
        try:
            payload = _build_eru(args)
        except BuildError as exc:
            print(f"ERROR: {exc}", file=sys.stderr)
            return 1
        _write_payload(args.output, payload)
        return 0

    snapshots: Dict[str, Dict[str, Dict[str, PricePoint]]]
    snapshot_sources: Dict[str, str]
    source_files: Dict[str, SourceMeta]
    try:
        snapshots, snapshot_sources, source_files = _build_snapshot_map(args.files)
        _snapshot_coverage_complete(snapshots)
        _check_price_movement(snapshots, args.allow_large_moves)
    except BuildError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    try:
        payload = _build_payload(snapshots, source_files)
    except BuildError as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        return 1

    _write_payload(args.output, payload)
    return 0


def main(argv: Optional[List[str]] = None) -> int:
    return build(argv)


if __name__ == "__main__":
    raise SystemExit(main())
